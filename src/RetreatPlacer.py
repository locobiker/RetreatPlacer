#!/usr/bin/env python3
"""
Retreat Center Room Placement using Google OR-Tools CP-SAT Solver.

Usage:
    pip install ortools openpyxl pandas
    python RetreatPlacer.py [RoomMap.xlsx] [PeopleToPlace.xlsx] [FilledRoomMap.xlsx]

Inputs:
  - RoomMap.xlsx:       BuildingName, RoomName, RoomFloor (1|2), #BottomBunk, #TopBunk
  - PeopleToPlace.xlsx: FirstName, LastName, OrgName, GroupName, AttachName,
                        RoomLocationPref (1|Any), BunkPref (Bottom|Any)

Output:
  - FilledRoomMap.xlsx: BuildingName, RoomName, FirstName, LastName, OrgName, GroupName, RoomFloor, Bunk
                        (includes an "Unplaced" sheet with people who couldn't be assigned)

Rules (hard constraints):
  - BunkPref: Bottom => must get a bottom bunk (accessibility)
  - RoomLocationPref: 1 => must be on floor 1 (accessibility)
  - Room capacity cannot be exceeded
  - AttachName pairs must share the same room

Rules (soft constraints, optimized):
  - GroupName members should share the same room(s)
  - OrgName members should share the same building(s)
  - Maximize total people placed
"""

import sys
import pandas as pd
from collections import defaultdict
from difflib import SequenceMatcher

from ortools.sat.python import cp_model

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def load_data(room_file, people_file):
    rooms_df = pd.read_excel(room_file)
    rooms_df.columns = rooms_df.columns.str.strip()

    people_df = pd.read_excel(people_file)
    people_df.columns = people_df.columns.str.strip()

    rooms_df['RoomFloor'] = rooms_df['RoomFloor'].astype(int)
    rooms_df['#BottomBunk'] = rooms_df['#BottomBunk'].astype(int)
    rooms_df['#TopBunk'] = rooms_df['#TopBunk'].astype(int)

    for col in ['FirstName', 'LastName', 'OrgName', 'GroupName', 'AttachName',
                'RoomLocationPref', 'BunkPref']:
        people_df[col] = (people_df[col].fillna('').astype(str).str.strip()
                          .replace(['nan', 'NaN', 'None', 'none'], ''))

    # Normalize GroupName and OrgName to handle case inconsistencies
    # (e.g. "Momlife" vs "MomLife" should be treated as the same group).
    # Strategy: use lowercase as the key, first occurrence as the canonical display name.
    for col in ['GroupName', 'OrgName']:
        canonical = {}  # lowercase -> first-seen display form
        for val in people_df[col]:
            if val and val.lower() not in canonical:
                canonical[val.lower()] = val
        people_df[col] = people_df[col].apply(
            lambda v: canonical.get(v.lower(), v) if v else v)

    # Auto-assign GroupName when AttachName references a known group.
    # Some people put their group name in AttachName instead of GroupName
    # (e.g. Hannah Emerson has AttachName="Momlife" but empty GroupName).
    known_groups = set(people_df['GroupName'].unique()) - {'', 'nan', 'NaN'}
    group_lower_map = {g.lower().replace(' ', ''): g for g in known_groups}
    for idx, row in people_df.iterrows():
        att = row['AttachName'].lower().replace(' ', '')
        if att and not row['GroupName'] and att in group_lower_map:
            people_df.at[idx, 'GroupName'] = group_lower_map[att]
            print(f"  Auto-assigned GroupName='{group_lower_map[att]}' for "
                  f"{row['FirstName']} {row['LastName']} (AttachName was '{row['AttachName']}')")

    return rooms_df, people_df


# ---------------------------------------------------------------------------
# Fuzzy AttachName resolution
# ---------------------------------------------------------------------------

def resolve_attach_names(people_df):
    """
    Resolve AttachName references using fuzzy matching with org/group affinity.

    Many AttachName values in real data have typos, case differences, nicknames,
    or reference non-person concepts (group names, etc.). This function:
    1. Tries exact match (case-insensitive)
    2. Tries common nickname expansions (e.g. Jess -> Jessica)
    3. Uses last-name, first-name, and prefix matching with org/group tiebreaking
    4. Falls back to fuzzy string matching boosted by org/group affinity
    5. Identifies non-person references (group names, org names, etc.)

    When multiple candidates match at a given step, candidates sharing the
    source person's OrgName or GroupName are preferred. This prevents false
    matches across organizational boundaries (e.g. "Heather" resolving to the
    Heather in the same org rather than one in a different org).

    Returns:
        resolved_attach: dict mapping person index -> attached person index
        attach_warnings: list of (person_name, attach_value, warning_msg)
    """
    people = people_df.to_dict('records')

    # Build lookup: lowercase full name -> person index
    name_to_idx = {}
    for i, p in enumerate(people):
        fn, ln = p['FirstName'], p['LastName']
        if fn and ln:
            name_to_idx[f"{fn} {ln}".lower()] = i

    # Also build last-name lookup for partial matches
    lastname_to_indices = defaultdict(list)
    for i, p in enumerate(people):
        if p['LastName']:
            lastname_to_indices[p['LastName'].lower()].append(i)

    # Common nickname mappings
    nicknames = {
        'jess': 'jessica', 'jen': 'jennifer', 'mike': 'michael',
        'chris': 'christina', 'liz': 'elizabeth', 'bob': 'robert',
        'bill': 'william', 'sam': 'samantha', 'dan': 'daniel',
        'nick': 'nicholas', 'nicki': 'nicole', 'hanna': 'hannah',
        'stacey': 'stacey', 'stacy': 'stacey', 'sherri': 'sheri',
        'nikki': 'nicole', 'cathy': 'catherine', 'kami': 'kameron',
    }

    def _affinity_score(source_idx, candidate_idx):
        """
        Return a numeric affinity bonus for how closely two people are related
        organizationally. Higher = stronger match preference.
          +2  same GroupName (non-empty)
          +1  same OrgName (non-empty)
           0  no organizational relationship
        """
        src, cand = people[source_idx], people[candidate_idx]
        score = 0
        if src['GroupName'] and src['GroupName'] == cand['GroupName']:
            score += 2
        if src['OrgName'] and src['OrgName'] == cand['OrgName']:
            score += 1
        return score

    def _pick_best_candidate(source_idx, candidates):
        """
        From a list of candidate indices, pick the one with the highest
        org/group affinity to source. If tied with zero affinity, return
        None (ambiguous). If only one candidate, return it directly.
        """
        if len(candidates) == 1:
            return candidates[0]
        if not candidates:
            return None

        scored = [(c, _affinity_score(source_idx, c)) for c in candidates]
        scored.sort(key=lambda x: -x[1])

        best_aff = scored[0][1]
        top_tier = [c for c, a in scored if a == best_aff]

        if len(top_tier) == 1:
            return top_tier[0]
        if best_aff > 0:
            return scored[0][0]
        return None

    # Build non-person reference set dynamically from actual group/org names,
    # plus known patterns. This catches AttachName values like "MomLife",
    # "Rock Point", "CelebrateRecovery" that reference groups/orgs, not people.
    non_person_exact = {
        '30/40s', '30s 40s',
        'young ladies',
        'rock point people',
    }
    # Add all actual GroupName and OrgName values (lowercased, with/without spaces)
    for col in ['GroupName', 'OrgName']:
        for val in set(people_df[col].unique()):
            if val and val not in ('', 'nan'):
                non_person_exact.add(val.lower())
                non_person_exact.add(val.lower().replace(' ', ''))

    non_person_prefixes = [
        'cr - ',
    ]
    non_person_contains = [
        ', ',
        ' and ',
    ]

    resolved_attach = {}
    attach_warnings = []
    rejected_matches = defaultdict(set)  # person_idx -> set of rejected candidate indices

    for i, p in enumerate(people):
        att = p['AttachName']
        if not att:
            continue

        att_lower = att.lower()

        # Skip known non-person references
        is_non_person = (
            att_lower in non_person_exact or
            any(att_lower.startswith(pfx) for pfx in non_person_prefixes) or
            any(pat in att_lower for pat in non_person_contains)
        )
        if is_non_person:
            attach_warnings.append((
                f"{p['FirstName']} {p['LastName']}", att,
                f"Skipped: '{att}' appears to be a group/org reference, not a person name"
            ))
            continue

        # 1. Exact case-insensitive match (unambiguous — no affinity needed)
        if att_lower in name_to_idx:
            target = name_to_idx[att_lower]
            if target != i:
                resolved_attach[i] = target
                continue

        # 2. Try nickname expansion (single lookup — unambiguous if found)
        parts = att_lower.split()
        if len(parts) == 2:
            first, last = parts
            expanded_first = nicknames.get(first, first)
            expanded_key = f"{expanded_first} {last}"
            if expanded_key in name_to_idx:
                target = name_to_idx[expanded_key]
                if target != i:
                    resolved_attach[i] = target
                    continue

            # Try matching on last name — use affinity to disambiguate
            if last in lastname_to_indices:
                candidates = [c for c in lastname_to_indices[last] if c != i]
                best = _pick_best_candidate(i, candidates)
                if best is not None:
                    # Verify first name is at least plausible (not completely different)
                    cand_first = people[best]['FirstName'].lower()
                    first_sim = SequenceMatcher(None, first, cand_first).ratio()
                    aff = _affinity_score(i, best)
                    # Accept if: first names are similar (>0.4), OR same org/group,
                    # OR the attach first name is a known nickname of the candidate
                    if first_sim >= 0.4 or aff > 0 or nicknames.get(first, first) == cand_first:
                        resolved_attach[i] = best
                        target_name = f"{people[best]['FirstName']} {people[best]['LastName']}"
                        if len(candidates) > 1:
                            attach_warnings.append((
                                f"{p['FirstName']} {p['LastName']}", att,
                                f"Last-name matched to '{target_name}' "
                                f"(picked from {len(candidates)} candidates, affinity={aff})"
                            ))
                        continue
                    else:
                        attach_warnings.append((
                            f"{p['FirstName']} {p['LastName']}", att,
                            f"Last-name match rejected: '{first}' vs '{cand_first}' "
                            f"(sim={first_sim:.2f}, affinity={aff}) — likely different person"
                        ))
                        # Track rejected candidates so fuzzy step doesn't re-match them
                        rejected_matches[i].add(best)

        # 3. Try first-name-only match with affinity tiebreaking
        if len(parts) == 1:
            first = parts[0]
            candidates = [idx for name_key, idx in name_to_idx.items()
                          if idx != i and name_key.split()[0] == first]
            best = _pick_best_candidate(i, candidates)
            if best is not None:
                resolved_attach[i] = best
                target_name = f"{people[best]['FirstName']} {people[best]['LastName']}"
                aff = _affinity_score(i, best)
                attach_warnings.append((
                    f"{p['FirstName']} {p['LastName']}", att,
                    f"First-name matched to '{target_name}' "
                    f"(from {len(candidates)} candidates, affinity={aff})"
                ))
                continue

        # 3b. Try prefix matching with affinity tiebreaking
        if len(parts) == 2:
            first, last_prefix = parts
            if len(last_prefix) <= 3:
                candidates = [idx for name_key, idx in name_to_idx.items()
                              if idx != i and len(name_key.split()) == 2
                              and name_key.split()[0] == first
                              and name_key.split()[1].startswith(last_prefix)]
                best = _pick_best_candidate(i, candidates)
                if best is not None:
                    resolved_attach[i] = best
                    target_name = f"{people[best]['FirstName']} {people[best]['LastName']}"
                    attach_warnings.append((
                        f"{p['FirstName']} {p['LastName']}", att,
                        f"Prefix matched to '{target_name}'"
                    ))
                    continue

        # 4. Fuzzy match with org/group affinity boost
        #
        # The affinity bonus (0.15 per point) means a same-org candidate at
        # raw=0.60 beats a stranger at 0.72. People are far more likely to
        # reference someone in their own org by nickname or typo.
        AFFINITY_BOOST = 0.15

        best_combined = 0
        best_raw = 0
        best_idx = None
        rejected = rejected_matches.get(i, set())
        for name_key, idx in name_to_idx.items():
            if idx == i or idx in rejected:
                continue
            raw = SequenceMatcher(None, att_lower, name_key).ratio()
            aff = _affinity_score(i, idx)
            combined = raw + aff * AFFINITY_BOOST
            if combined > best_combined or (combined == best_combined and raw > best_raw):
                best_combined = combined
                best_raw = raw
                best_idx = idx

        # Require minimum raw score of 0.55 (with affinity) or 0.70 (without)
        aff = _affinity_score(i, best_idx) if best_idx is not None else 0
        min_threshold = 0.60 if aff > 0 else 0.70

        if best_raw >= min_threshold and best_idx is not None:
            resolved_attach[i] = best_idx
            target_name = f"{people[best_idx]['FirstName']} {people[best_idx]['LastName']}"
            if best_raw < 0.95:
                attach_warnings.append((
                    f"{p['FirstName']} {p['LastName']}", att,
                    f"Fuzzy matched to '{target_name}' "
                    f"(raw={best_raw:.2f}, affinity={aff}, combined={best_combined:.2f})"
                ))
            continue

        # 5. Unresolved
        attach_warnings.append((
            f"{p['FirstName']} {p['LastName']}", att,
            f"UNRESOLVED: No match found for '{att}' "
            f"(best raw={best_raw:.2f}, affinity={aff})"
        ))

    return resolved_attach, attach_warnings


# ---------------------------------------------------------------------------
# Slot construction
# ---------------------------------------------------------------------------

def build_slots(rooms_df):
    """
    Expand rooms into individual bed-slots.
    Returns:
        slots     - list of (building, room, floor, bunk_type)
        room_keys - ordered list of unique (building, room) tuples
    """
    slots = []
    for _, row in rooms_df.iterrows():
        bldg, room, floor = row['BuildingName'], row['RoomName'], int(row['RoomFloor'])
        for _ in range(int(row['#BottomBunk'])):
            slots.append((bldg, room, floor, 'Bottom'))
        for _ in range(int(row['#TopBunk'])):
            slots.append((bldg, room, floor, 'Top'))

    room_keys = list(dict.fromkeys((s[0], s[1]) for s in slots))
    return slots, room_keys


# ---------------------------------------------------------------------------
# Pre-assignment: compute ideal building assignments for orgs
# ---------------------------------------------------------------------------

def compute_org_building_affinity(rooms_df, people_df):
    """
    Pre-compute which buildings each org should be assigned to, based on
    org size and building capacity. This helps the solver by providing
    explicit building targets rather than relying purely on pairwise constraints.

    Returns:
        org_building_map: dict of org_name -> set of preferred building names
    """
    # Compute building capacities
    bldg_capacity = {}
    for _, row in rooms_df.iterrows():
        bldg = row['BuildingName']
        cap = int(row['#BottomBunk']) + int(row['#TopBunk'])
        bldg_capacity[bldg] = bldg_capacity.get(bldg, 0) + cap

    # Compute org sizes
    org_sizes = defaultdict(int)
    for _, row in people_df.iterrows():
        org = row['OrgName']
        if org:
            org_sizes[org] += 1

    # Sort orgs by size descending (largest first for greedy assignment)
    sorted_orgs = sorted(org_sizes.items(), key=lambda x: -x[1])
    # Sort buildings by capacity descending
    sorted_bldgs = sorted(bldg_capacity.items(), key=lambda x: -x[1])

    org_building_map = {}
    remaining_capacity = dict(bldg_capacity)

    for org, size in sorted_orgs:
        # Greedy: assign this org to buildings with most remaining capacity
        # Try to fit in as few buildings as possible
        assigned_bldgs = set()
        needed = size

        # Sort remaining buildings by capacity desc
        candidates = sorted(remaining_capacity.items(), key=lambda x: -x[1])

        for bldg, cap in candidates:
            if needed <= 0:
                break
            if cap > 0:
                assigned_bldgs.add(bldg)
                take = min(cap, needed)
                remaining_capacity[bldg] -= take
                needed -= take

        org_building_map[org] = assigned_bldgs

    return org_building_map


# ---------------------------------------------------------------------------
# Solver
# ---------------------------------------------------------------------------

def solve_placement(rooms_df, people_df):
    """
    Build and solve the CP-SAT model for room placement.

    Uses room-level assignment (not individual bunk slots) for efficiency.
    Each person is assigned to a room, then bunk types are assigned in
    post-processing based on BunkPref and room bunk counts.
    """
    slots, room_keys = build_slots(rooms_df)

    people = people_df.to_dict('records')
    num_people = len(people)

    # Resolve AttachNames with fuzzy matching
    resolved_attach, attach_warnings = resolve_attach_names(people_df)

    print("\n--- AttachName Resolution ---")
    for person_name, att_val, msg in attach_warnings:
        print(f"  {person_name}: {msg}")
    print(f"  Total resolved: {len(resolved_attach)} pairs")

    # Room / building lookups
    num_rooms = len(room_keys)
    UNASSIGNED_ROOM = num_rooms

    room_to_id = {rk: i for i, rk in enumerate(room_keys)}
    bldg_names = list(dict.fromkeys(rk[0] for rk in room_keys))
    bldg_to_id = {b: i for i, b in enumerate(bldg_names)}
    UNASSIGNED_BLDG = len(bldg_names)

    # Room properties
    room_floor = {}
    room_bldg_id = {}
    room_bottom_cap = {}
    room_total_cap = {}
    for _, row in rooms_df.iterrows():
        rk = (row['BuildingName'], row['RoomName'])
        rid = room_to_id[rk]
        room_floor[rid] = int(row['RoomFloor'])
        room_bldg_id[rid] = bldg_to_id[row['BuildingName']]
        room_bottom_cap[rid] = int(row['#BottomBunk'])
        room_total_cap[rid] = int(row['#BottomBunk']) + int(row['#TopBunk'])

    # Org-building affinity
    org_bldg_map = compute_org_building_affinity(rooms_df, people_df)
    print("\n--- Org-Building Affinity (pre-computed) ---")
    for org, bldgs in org_bldg_map.items():
        print(f"  {org} ({sum(1 for p in people if p['OrgName']==org)} people) -> {bldgs}")

    # Array for Element constraint: room_id -> bldg_id
    room_bldg_array = [room_bldg_id.get(r, UNASSIGNED_BLDG) for r in range(num_rooms)]
    room_bldg_array.append(UNASSIGNED_BLDG)

    # -----------------------------------------------------------------------
    model = cp_model.CpModel()

    # Decision variable: room_id[p] = which room (or UNASSIGNED_ROOM)
    room_id = [model.NewIntVar(0, UNASSIGNED_ROOM, f'rid_{p}') for p in range(num_people)]

    # Derived: bldg_id[p]
    bldg_id = []
    for p in range(num_people):
        b = model.NewIntVar(0, UNASSIGNED_BLDG, f'bid_{p}')
        model.AddElement(room_id[p], room_bldg_array, b)
        bldg_id.append(b)

    # Boolean: is person p assigned?
    assigned = []
    for p in range(num_people):
        b = model.NewBoolVar(f'asgn_{p}')
        model.Add(room_id[p] != UNASSIGNED_ROOM).OnlyEnforceIf(b)
        model.Add(room_id[p] == UNASSIGNED_ROOM).OnlyEnforceIf(b.Not())
        assigned.append(b)

    # Per-person, per-room indicator: in_room[p][rid]
    in_room = {}
    for rid in range(num_rooms):
        for p in range(num_people):
            b = model.NewBoolVar(f'ir_{p}_{rid}')
            model.Add(room_id[p] == rid).OnlyEnforceIf(b)
            model.Add(room_id[p] != rid).OnlyEnforceIf(b.Not())
            in_room[(p, rid)] = b

    # ===================== HARD CONSTRAINTS =====================

    # H1: Room total capacity
    for rid in range(num_rooms):
        people_in_room = [in_room[(p, rid)] for p in range(num_people)]
        model.Add(sum(people_in_room) <= room_total_cap[rid])

    # H2: Bottom bunk capacity per room
    bottom_needers = [p for p, person in enumerate(people)
                      if person['BunkPref'].lower() == 'bottom']
    for rid in range(num_rooms):
        bottom_in_room = [in_room[(p, rid)] for p in bottom_needers]
        if bottom_in_room:
            model.Add(sum(bottom_in_room) <= room_bottom_cap[rid])

    # H3: Floor preference
    for p, person in enumerate(people):
        if person['RoomLocationPref'] == '1':
            allowed = [rid for rid in range(num_rooms) if room_floor[rid] == 1]
            allowed.append(UNASSIGNED_ROOM)
            model.AddAllowedAssignments([room_id[p]], [[v] for v in allowed])

    # H4 / S-Attach: mutual pairs = HARD, one-directional = SOFT
    mutual_pairs = set()
    onedir_pairs = set()
    for p1, p2 in resolved_attach.items():
        pair = tuple(sorted([p1, p2]))
        if resolved_attach.get(p2) == p1:
            mutual_pairs.add(pair)
        else:
            onedir_pairs.add(pair)

    print(f"\n  Attach pairs: {len(mutual_pairs)} mutual (HARD), "
          f"{len(onedir_pairs)} one-directional (SOFT)")

    for p1, p2 in mutual_pairs:
        both = model.NewBoolVar(f'att_b_{p1}_{p2}')
        model.AddBoolAnd([assigned[p1], assigned[p2]]).OnlyEnforceIf(both)
        model.AddBoolOr([assigned[p1].Not(), assigned[p2].Not()]).OnlyEnforceIf(both.Not())
        model.Add(room_id[p1] == room_id[p2]).OnlyEnforceIf(both)

    attach_soft_vars = []
    attach_soft_mismatch = []
    for p1, p2 in onedir_pairs:
        both = model.NewBoolVar(f'ats_b_{p1}_{p2}')
        model.AddBoolAnd([assigned[p1], assigned[p2]]).OnlyEnforceIf(both)
        model.AddBoolOr([assigned[p1].Not(), assigned[p2].Not()]).OnlyEnforceIf(both.Not())

        same = model.NewBoolVar(f'ats_s_{p1}_{p2}')
        model.Add(room_id[p1] == room_id[p2]).OnlyEnforceIf(same)
        model.Add(room_id[p1] != room_id[p2]).OnlyEnforceIf(same.Not())

        ok = model.NewBoolVar(f'ats_ok_{p1}_{p2}')
        model.AddBoolAnd([both, same]).OnlyEnforceIf(ok)
        model.AddBoolOr([both.Not(), same.Not()]).OnlyEnforceIf(ok.Not())
        attach_soft_vars.append(ok)

        mis = model.NewBoolVar(f'ats_mis_{p1}_{p2}')
        model.AddBoolAnd([both, same.Not()]).OnlyEnforceIf(mis)
        model.AddBoolOr([both.Not(), same]).OnlyEnforceIf(mis.Not())
        attach_soft_mismatch.append(mis)

    # ===================== SOFT CONSTRAINTS =====================

    # S1: GroupName same room
    groups = defaultdict(list)
    for p, person in enumerate(people):
        if person['GroupName']:
            groups[person['GroupName']].append(p)

    group_vars, group_mismatch_vars = [], []
    for gname, members in groups.items():
        for i in range(len(members) - 1):
            p1, p2 = members[i], members[i + 1]
            both = model.NewBoolVar(f'gb_{gname}_{i}')
            model.AddBoolAnd([assigned[p1], assigned[p2]]).OnlyEnforceIf(both)
            model.AddBoolOr([assigned[p1].Not(), assigned[p2].Not()]).OnlyEnforceIf(both.Not())
            same = model.NewBoolVar(f'gs_{gname}_{i}')
            model.Add(room_id[p1] == room_id[p2]).OnlyEnforceIf(same)
            model.Add(room_id[p1] != room_id[p2]).OnlyEnforceIf(same.Not())
            ok = model.NewBoolVar(f'gok_{gname}_{i}')
            model.AddBoolAnd([both, same]).OnlyEnforceIf(ok)
            model.AddBoolOr([both.Not(), same.Not()]).OnlyEnforceIf(ok.Not())
            group_vars.append(ok)
            mismatch = model.NewBoolVar(f'gmis_{gname}_{i}')
            model.AddBoolAnd([both, same.Not()]).OnlyEnforceIf(mismatch)
            model.AddBoolOr([both.Not(), same]).OnlyEnforceIf(mismatch.Not())
            group_mismatch_vars.append(mismatch)

    # S2: OrgName same building
    orgs = defaultdict(list)
    for p, person in enumerate(people):
        if person['OrgName']:
            orgs[person['OrgName']].append(p)

    org_vars, org_mismatch_vars = [], []
    for oname, members in orgs.items():
        for i in range(len(members) - 1):
            p1, p2 = members[i], members[i + 1]
            both = model.NewBoolVar(f'ob_{oname}_{i}')
            model.AddBoolAnd([assigned[p1], assigned[p2]]).OnlyEnforceIf(both)
            model.AddBoolOr([assigned[p1].Not(), assigned[p2].Not()]).OnlyEnforceIf(both.Not())
            same = model.NewBoolVar(f'os_{oname}_{i}')
            model.Add(bldg_id[p1] == bldg_id[p2]).OnlyEnforceIf(same)
            model.Add(bldg_id[p1] != bldg_id[p2]).OnlyEnforceIf(same.Not())
            ok = model.NewBoolVar(f'ook_{oname}_{i}')
            model.AddBoolAnd([both, same]).OnlyEnforceIf(ok)
            model.AddBoolOr([both.Not(), same.Not()]).OnlyEnforceIf(ok.Not())
            org_vars.append(ok)
            mismatch = model.NewBoolVar(f'omis_{oname}_{i}')
            model.AddBoolAnd([both, same.Not()]).OnlyEnforceIf(mismatch)
            model.AddBoolOr([both.Not(), same]).OnlyEnforceIf(mismatch.Not())
            org_mismatch_vars.append(mismatch)

    # S3: Org-building affinity
    affinity_vars = []
    for p, person in enumerate(people):
        org = person['OrgName']
        if org and org in org_bldg_map:
            pref_bldg_ids = [bldg_to_id[b] for b in org_bldg_map[org] if b in bldg_to_id]
            pref_bldg_ids.append(UNASSIGNED_BLDG)
            in_pref_bldg = model.NewBoolVar(f'ipb_{p}')
            model.AddAllowedAssignments([bldg_id[p]], [[v] for v in pref_bldg_ids]).OnlyEnforceIf(in_pref_bldg)
            model.AddForbiddenAssignments([bldg_id[p]], [[v] for v in pref_bldg_ids]).OnlyEnforceIf(in_pref_bldg.Not())
            in_pref = model.NewBoolVar(f'afn_{p}')
            model.AddBoolAnd([assigned[p], in_pref_bldg]).OnlyEnforceIf(in_pref)
            model.AddBoolOr([assigned[p].Not(), in_pref_bldg.Not()]).OnlyEnforceIf(in_pref.Not())
            affinity_vars.append(in_pref)

    # ===================== OBJECTIVE =====================
    W_PLACE    = 10000
    W_GROUP    = 1000
    W_ATTACH   = 800
    W_ORG      = 100
    W_AFFINITY = 200

    model.Maximize(
        sum(W_PLACE * a for a in assigned) +
        sum(W_GROUP * g for g in group_vars) -
        sum(W_GROUP * m for m in group_mismatch_vars) +
        sum(W_ATTACH * a for a in attach_soft_vars) -
        sum(W_ATTACH * m for m in attach_soft_mismatch) +
        sum(W_ORG   * o for o in org_vars) -
        sum(W_ORG   * m for m in org_mismatch_vars) +
        sum(W_AFFINITY * a for a in affinity_vars)
    )

    # ===================== SOLVE =====================
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 300
    solver.parameters.num_search_workers = 8
    status = solver.Solve(model)

    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        print("ERROR: Solver found no feasible solution!")
        return None, [(p, ["Solver failure"], '') for p in people], slots, attach_warnings, resolved_attach

    # ===================== EXTRACT RESULTS =====================
    attach_resolved_names = {}
    for p_idx, target_idx in resolved_attach.items():
        t = people[target_idx]
        attach_resolved_names[p_idx] = f"{t['FirstName']} {t['LastName']}"

    results = []
    unplaced = []

    # Group placed people by room for bunk post-processing
    room_assignments = defaultdict(list)
    for p, person in enumerate(people):
        rid = solver.Value(room_id[p])
        if rid == UNASSIGNED_ROOM:
            reasons = _diagnose_room(person, p, people, resolved_attach,
                                     room_bottom_cap, room_floor, num_rooms,
                                     solver, assigned, room_id, bldg_id)
            attach_res = attach_resolved_names.get(p, '')
            unplaced.append((person, reasons, attach_res))
        else:
            room_assignments[rid].append((p, person))

    for rid, members in room_assignments.items():
        bldg_name, room_name = room_keys[rid]
        floor = room_floor[rid]
        bcap = room_bottom_cap[rid]

        # Assign bunks: bottom-needers first, then fill remaining bottom, then top
        bottom_first = sorted(members, key=lambda x: (0 if x[1]['BunkPref'].lower() == 'bottom' else 1))
        bottom_used = 0
        for p, person in bottom_first:
            if bottom_used < bcap:
                btype = 'Bottom'
                bottom_used += 1
            else:
                btype = 'Top'
            results.append({
                'BuildingName': bldg_name,
                'RoomName': room_name,
                'FirstName': person['FirstName'],
                'LastName': person['LastName'],
                'OrgName': person['OrgName'],
                'GroupName': person['GroupName'],
                'RoomFloor': floor,
                'Bunk': btype,
                'AttachName': person['AttachName'],
                'AttachResolved': attach_resolved_names.get(p, ''),
            })

    # Report
    group_sat  = sum(solver.Value(v) for v in group_vars)
    group_mis  = sum(solver.Value(v) for v in group_mismatch_vars)
    attach_sat = sum(solver.Value(v) for v in attach_soft_vars)
    attach_mis = sum(solver.Value(v) for v in attach_soft_mismatch)
    org_sat    = sum(solver.Value(v) for v in org_vars)
    org_mis    = sum(solver.Value(v) for v in org_mismatch_vars)
    affinity_sat = sum(solver.Value(v) for v in affinity_vars)
    print(f"\nSoft constraint satisfaction:")
    print(f"  Group-same-room:      {group_sat}/{len(group_vars)} matched, {group_mis} mismatched")
    print(f"  Attach-same-room:     {attach_sat}/{len(attach_soft_vars)} matched, {attach_mis} mismatched")
    print(f"  Mutual attach pairs:  {len(mutual_pairs)} (hard)")
    print(f"  Org-same-building:    {org_sat}/{len(org_vars)} matched, {org_mis} mismatched")
    print(f"  Org-building affinity: {affinity_sat}/{len(affinity_vars)} in preferred building")
    print(f"  Solution: {'OPTIMAL' if status == cp_model.OPTIMAL else 'FEASIBLE'}")

    return results, unplaced, slots, attach_warnings, resolved_attach

def _diagnose_room(person, p_idx, people, resolved_attach,
                   room_bottom_cap, room_floor_map, num_rooms,
                   solver, assigned, room_id, bldg_id):
    """Generate human-readable reasons why a person could not be placed."""
    reasons = []
    needs_bottom = person['BunkPref'].lower() == 'bottom'
    needs_floor1 = person['RoomLocationPref'] == '1'

    floor1_rooms = sum(1 for r in range(num_rooms) if room_floor_map.get(r) == 1)
    bottom_total = sum(room_bottom_cap.get(r, 0) for r in range(num_rooms))
    floor1_bottom = sum(room_bottom_cap.get(r, 0) for r in range(num_rooms)
                        if room_floor_map.get(r) == 1)

    if needs_bottom and needs_floor1:
        reasons.append(f"Needs bottom bunk on floor 1 ({floor1_bottom} such bunks exist, likely full)")
    elif needs_bottom:
        reasons.append(f"Needs bottom bunk ({bottom_total} exist total, high demand)")
    elif needs_floor1:
        reasons.append(f"Needs floor 1 ({floor1_rooms} rooms exist)")

    # Check attachment
    if p_idx in resolved_attach:
        partner_idx = resolved_attach[p_idx]
        partner = people[partner_idx]
        partner_name = f"{partner['FirstName']} {partner['LastName']}"
        partner_placed = solver.Value(assigned[partner_idx])
        if not partner_placed:
            reasons.append(f"Attached to '{partner_name}' who is also unplaced")
        else:
            reasons.append(f"Attached to '{partner_name}' (placed) — room may have been full")

    att = person['AttachName']
    if att and p_idx not in resolved_attach:
        reasons.append(f"AttachName '{att}' could not be resolved to a person in the list")

    grp = person['GroupName']
    if grp:
        reasons.append(f"Group '{grp}' cohesion constraints may have limited options")

    org = person['OrgName']
    if org:
        reasons.append(f"Org '{org}' building affinity may have limited available slots")

    if not reasons:
        reasons.append("Capacity exhausted or competing constraints")
    return reasons


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

def write_output(results, unplaced, output_file, attach_warnings=None, resolved_attach=None):
    wb = Workbook()

    # ===== Sheet 1: FilledRoomMap =====
    ws = wb.active
    ws.title = "FilledRoomMap"

    headers = ['BuildingName', 'RoomName', 'FirstName', 'LastName', 'OrgName',
               'GroupName', 'RoomFloor', 'Bunk', 'AttachName', 'AttachResolved']

    hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    hdr_fill = PatternFill('solid', fgColor='4472C4')
    hdr_align = Alignment(horizontal='center', vertical='center')
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font, cell.fill, cell.alignment, cell.border = hdr_font, hdr_fill, hdr_align, thin

    # Sort by building > room > bunk(Bottom first)
    results.sort(key=lambda r: (r['BuildingName'], r['RoomName'],
                                 0 if r['Bunk'] == 'Bottom' else 1))

    alt = PatternFill('solid', fgColor='D9E2F3')
    dfont = Font(name='Arial', size=11)

    for i, rd in enumerate(results, 2):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=i, column=c, value=rd[h])
            cell.font = dfont
            cell.border = thin
            cell.alignment = Alignment(
                horizontal='center' if h in ('RoomFloor', 'Bunk') else 'left')
            if i % 2 == 0:
                cell.fill = alt

    widths = {'A': 18, 'B': 16, 'C': 16, 'D': 16, 'E': 18, 'F': 18, 'G': 12, 'H': 10,
              'I': 22, 'J': 22}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.auto_filter.ref = f"A1:J{len(results) + 1}"

    # ===== Sheet 2: Unplaced People =====
    ws2 = wb.create_sheet("Unplaced")

    unplaced_headers = ['FirstName', 'LastName', 'OrgName', 'GroupName', 'AttachName',
                        'AttachResolved', 'RoomLocationPref', 'BunkPref', 'Reasons']
    unplaced_fill = PatternFill('solid', fgColor='C00000')
    unplaced_hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)

    for c, h in enumerate(unplaced_headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = unplaced_hdr_font
        cell.fill = unplaced_fill
        cell.alignment = hdr_align
        cell.border = thin

    light_red = PatternFill('solid', fgColor='FFC7CE')
    for i, (person, reasons, attach_res) in enumerate(unplaced, 2):
        vals = [person.get('FirstName', ''), person.get('LastName', ''),
                person.get('OrgName', ''), person.get('GroupName', ''),
                person.get('AttachName', ''), attach_res,
                person.get('RoomLocationPref', ''), person.get('BunkPref', ''),
                '; '.join(reasons)]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row=i, column=c, value=v)
            cell.font = dfont
            cell.border = thin
            if i % 2 == 0:
                cell.fill = light_red

    unplaced_widths = {'A': 14, 'B': 16, 'C': 16, 'D': 18, 'E': 22,
                       'F': 22, 'G': 18, 'H': 12, 'I': 60}
    for col, w in unplaced_widths.items():
        ws2.column_dimensions[col].width = w
    if unplaced:
        ws2.auto_filter.ref = f"A1:I{len(unplaced) + 1}"

    # ===== Sheet 3: Attach Name Warnings =====
    if attach_warnings:
        ws3 = wb.create_sheet("AttachWarnings")
        warn_headers = ['Person', 'AttachName Value', 'Resolution']
        warn_fill = PatternFill('solid', fgColor='ED7D31')
        for c, h in enumerate(warn_headers, 1):
            cell = ws3.cell(row=1, column=c, value=h)
            cell.font = unplaced_hdr_font
            cell.fill = warn_fill
            cell.alignment = hdr_align
            cell.border = thin

        for i, (person_name, att_val, msg) in enumerate(attach_warnings, 2):
            for c, v in enumerate([person_name, att_val, msg], 1):
                cell = ws3.cell(row=i, column=c, value=v)
                cell.font = dfont
                cell.border = thin

        ws3.column_dimensions['A'].width = 24
        ws3.column_dimensions['B'].width = 24
        ws3.column_dimensions['C'].width = 60

    # ===== Sheet 4: Summary =====
    ws4 = wb.create_sheet("Summary")
    ws4.column_dimensions['A'].width = 30
    ws4.column_dimensions['B'].width = 15
    ws4.column_dimensions['C'].width = 20

    title_font = Font(name='Arial', bold=True, size=14)
    section_font = Font(name='Arial', bold=True, size=12, color='4472C4')

    row = 1
    ws4.cell(row=row, column=1, value="Placement Summary").font = title_font
    row += 2

    ws4.cell(row=row, column=1, value="Total People").font = Font(name='Arial', bold=True, size=11)
    ws4.cell(row=row, column=2, value=len(results) + len(unplaced))
    row += 1
    ws4.cell(row=row, column=1, value="Placed").font = Font(name='Arial', bold=True, size=11, color='00B050')
    ws4.cell(row=row, column=2, value=len(results))
    row += 1
    ws4.cell(row=row, column=1, value="Unplaced").font = Font(name='Arial', bold=True, size=11, color='C00000')
    ws4.cell(row=row, column=2, value=len(unplaced))
    row += 2

    # Org breakdown
    ws4.cell(row=row, column=1, value="By Organization & Building").font = section_font
    row += 1
    ws4.cell(row=row, column=1, value="Organization").font = Font(name='Arial', bold=True, size=11)
    ws4.cell(row=row, column=2, value="Building").font = Font(name='Arial', bold=True, size=11)
    ws4.cell(row=row, column=3, value="Count").font = Font(name='Arial', bold=True, size=11)
    row += 1

    org_bldg_counts = defaultdict(lambda: defaultdict(int))
    for r in results:
        org_bldg_counts[r['OrgName']][r['BuildingName']] += 1

    for org in sorted(org_bldg_counts.keys()):
        for bldg in sorted(org_bldg_counts[org].keys()):
            ws4.cell(row=row, column=1, value=org).font = dfont
            ws4.cell(row=row, column=2, value=bldg).font = dfont
            ws4.cell(row=row, column=3, value=org_bldg_counts[org][bldg]).font = dfont
            row += 1

    wb.save(output_file)


# ---------------------------------------------------------------------------
# Debug report
# ---------------------------------------------------------------------------

def print_debug(results, unplaced, slots):
    print("\n" + "=" * 70)
    print("RETREAT CENTER PLACEMENT RESULTS")
    print("=" * 70)

    placed = len(results) if results else 0
    print(f"\n  Total bed slots : {len(slots)}")
    print(f"  People placed   : {placed}")
    print(f"  People unplaced : {len(unplaced)}")

    if results:
        by_bldg = defaultdict(int)
        by_room = defaultdict(int)
        for r in results:
            by_bldg[r['BuildingName']] += 1
            by_room[(r['BuildingName'], r['RoomName'])] += 1

        print("\n  By building:")
        for b, n in sorted(by_bldg.items()):
            print(f"    {b}: {n}")

        # Org-building crosstab
        org_bldg = defaultdict(lambda: defaultdict(int))
        for r in results:
            org_bldg[r['OrgName']][r['BuildingName']] += 1

        print("\n  Org-Building distribution:")
        for org in sorted(org_bldg.keys()):
            parts = [f"{b}:{n}" for b, n in sorted(org_bldg[org].items())]
            print(f"    {org}: {', '.join(parts)}")

    if unplaced:
        print("\n" + "-" * 70)
        print("UNPLACED PEOPLE")
        print("-" * 70)
        for person, reasons, attach_res in unplaced:
            print(f"\n  {person['FirstName']} {person['LastName']}  "
                  f"(Org={person['OrgName']}, Group={person['GroupName']}, "
                  f"Attach={person['AttachName']}, "
                  f"FloorPref={person['RoomLocationPref']}, "
                  f"BunkPref={person['BunkPref']})")
            for r in reasons:
                print(f"    -> {r}")
    else:
        print("\n  All people placed successfully!")

    print("=" * 70)


# ---------------------------------------------------------------------------
# Sample data generator (for testing)
# ---------------------------------------------------------------------------

def generate_sample_data():
    """Create sample RoomMap.xlsx and PeopleToPlace.xlsx for testing."""
    rooms = pd.DataFrame([
        {'BuildingName': 'Oak Lodge',   'RoomName': 'Room 101', 'RoomFloor': 1, '#BottomBunk': 2, '#TopBunk': 2},
        {'BuildingName': 'Oak Lodge',   'RoomName': 'Room 102', 'RoomFloor': 1, '#BottomBunk': 1, '#TopBunk': 1},
        {'BuildingName': 'Oak Lodge',   'RoomName': 'Room 201', 'RoomFloor': 2, '#BottomBunk': 2, '#TopBunk': 2},
        {'BuildingName': 'Oak Lodge',   'RoomName': 'Room 202', 'RoomFloor': 2, '#BottomBunk': 1, '#TopBunk': 1},
        {'BuildingName': 'Pine Hall',   'RoomName': 'Room A',   'RoomFloor': 1, '#BottomBunk': 3, '#TopBunk': 3},
        {'BuildingName': 'Pine Hall',   'RoomName': 'Room B',   'RoomFloor': 1, '#BottomBunk': 2, '#TopBunk': 2},
        {'BuildingName': 'Pine Hall',   'RoomName': 'Room C',   'RoomFloor': 2, '#BottomBunk': 2, '#TopBunk': 2},
        {'BuildingName': 'Maple House', 'RoomName': 'Suite 1',  'RoomFloor': 1, '#BottomBunk': 1, '#TopBunk': 0},
        {'BuildingName': 'Maple House', 'RoomName': 'Suite 2',  'RoomFloor': 1, '#BottomBunk': 1, '#TopBunk': 1},
    ])

    people = pd.DataFrame([
        {'FirstName': 'Alice',  'LastName': 'Smith',    'OrgName': 'Alpha', 'GroupName': 'Team1', 'AttachName': '',              'RoomLocationPref': 'Any', 'BunkPref': 'Any'},
        {'FirstName': 'Bob',    'LastName': 'Jones',    'OrgName': 'Alpha', 'GroupName': 'Team1', 'AttachName': '',              'RoomLocationPref': 'Any', 'BunkPref': 'Any'},
        {'FirstName': 'Carol',  'LastName': 'Davis',    'OrgName': 'Alpha', 'GroupName': 'Team1', 'AttachName': '',              'RoomLocationPref': 'Any', 'BunkPref': 'Any'},
        {'FirstName': 'Dave',   'LastName': 'Wilson',   'OrgName': 'Alpha', 'GroupName': 'Team2', 'AttachName': 'Eve Brown',     'RoomLocationPref': 'Any', 'BunkPref': 'Any'},
        {'FirstName': 'Eve',    'LastName': 'Brown',    'OrgName': 'Alpha', 'GroupName': 'Team2', 'AttachName': 'Dave Wilson',   'RoomLocationPref': 'Any', 'BunkPref': 'Any'},
        {'FirstName': 'Frank',  'LastName': 'Miller',   'OrgName': 'Beta',  'GroupName': 'Sales', 'AttachName': '',              'RoomLocationPref': '1',   'BunkPref': 'Bottom'},
        {'FirstName': 'Grace',  'LastName': 'Taylor',   'OrgName': 'Beta',  'GroupName': 'Sales', 'AttachName': '',              'RoomLocationPref': 'Any', 'BunkPref': 'Any'},
        {'FirstName': 'Hank',   'LastName': 'Anderson', 'OrgName': 'Beta',  'GroupName': 'Sales', 'AttachName': '',              'RoomLocationPref': 'Any', 'BunkPref': 'Any'},
        {'FirstName': 'Irene',  'LastName': 'Thomas',   'OrgName': 'Beta',  'GroupName': '',      'AttachName': '',              'RoomLocationPref': 'Any', 'BunkPref': 'Bottom'},
        {'FirstName': 'Jack',   'LastName': 'Moore',    'OrgName': 'Gamma', 'GroupName': '',      'AttachName': 'Karen White',   'RoomLocationPref': '1',   'BunkPref': 'Bottom'},
        {'FirstName': 'Karen',  'LastName': 'White',    'OrgName': 'Gamma', 'GroupName': '',      'AttachName': 'Jack Moore',    'RoomLocationPref': 'Any', 'BunkPref': 'Any'},
        {'FirstName': 'Leo',    'LastName': 'Harris',   'OrgName': 'Gamma', 'GroupName': 'Dev',   'AttachName': '',              'RoomLocationPref': 'Any', 'BunkPref': 'Any'},
        {'FirstName': 'Mona',   'LastName': 'Martin',   'OrgName': 'Gamma', 'GroupName': 'Dev',   'AttachName': '',              'RoomLocationPref': 'Any', 'BunkPref': 'Any'},
        {'FirstName': 'Nate',   'LastName': 'Garcia',   'OrgName': '',      'GroupName': '',      'AttachName': '',              'RoomLocationPref': 'Any', 'BunkPref': 'Any'},
        {'FirstName': 'Olivia', 'LastName': 'Martinez', 'OrgName': '',      'GroupName': '',      'AttachName': '',              'RoomLocationPref': '1',   'BunkPref': 'Bottom'},
    ])

    rooms.to_excel('RoomMap.xlsx', index=False)
    people.to_excel('PeopleToPlace.xlsx', index=False)
    print("Sample data generated: RoomMap.xlsx, PeopleToPlace.xlsx")
    print(f"  Capacity: {rooms['#BottomBunk'].sum() + rooms['#TopBunk'].sum()} slots")
    print(f"  People:   {len(people)}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    if '--generate-sample' in sys.argv:
        generate_sample_data()
        return

    room_file   = sys.argv[1] if len(sys.argv) > 1 else 'RoomMap.xlsx'
    people_file = sys.argv[2] if len(sys.argv) > 2 else 'PeopleToPlace.xlsx'
    output_file = sys.argv[3] if len(sys.argv) > 3 else 'FilledRoomMap.xlsx'

    print(f"Room map  : {room_file}")
    print(f"People    : {people_file}")
    print(f"Output    : {output_file}")

    # Pre-check: delete existing output file (catches Windows file-lock early)
    import os
    if os.path.exists(output_file):
        try:
            os.remove(output_file)
            print(f"  (removed existing {output_file})")
        except PermissionError:
            print(f"\nERROR: Cannot write to '{output_file}' — is it open in Excel?")
            print("Close the file and try again.")
            sys.exit(1)

    rooms_df, people_df = load_data(room_file, people_file)
    slots, _ = build_slots(rooms_df)

    print(f"\nRooms: {len(rooms_df)}  |  People: {len(people_df)}  |  "
          f"Slots: {len(slots)} "
          f"(Bottom: {sum(1 for s in slots if s[3]=='Bottom')}, "
          f"Top: {sum(1 for s in slots if s[3]=='Top')})")

    print("\nSolving with OR-Tools CP-SAT ...")
    results, unplaced, slots, attach_warnings, resolved_attach = solve_placement(rooms_df, people_df)

    print_debug(results, unplaced, slots)

    if results is not None:
        write_output(results, unplaced, output_file, attach_warnings, resolved_attach)
        print(f"\nOutput saved to: {output_file}")
        print(f"  Sheets: FilledRoomMap, Unplaced ({len(unplaced)} people), "
              f"AttachWarnings ({len(attach_warnings)} warnings), Summary")
    else:
        print("\nNo placements — no output file written.")


if __name__ == '__main__':
    main()
